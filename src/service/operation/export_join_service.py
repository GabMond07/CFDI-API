from src.database import db
from src.event_bus.publisher import publish_event
from typing import Any, Dict, Optional, Callable
import json
import pandas as pd
from dicttoxml import dicttoxml
import base64
import io
import logging
from datetime import datetime, timezone
from fastapi import HTTPException
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def flatten_dict(d: Dict, parent_key: str = '', sep: str = '_') -> Dict:
    """Aplana diccionarios anidados para CSV, Excel y otros formatos."""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep).items())
        elif isinstance(v, (list, tuple)):
            items.append((new_key, json.dumps(v)))  # Convertir listas a JSON string
        else:
            items.append((new_key, v))
    return dict(items)

def format_predefined_join_excel(data: list, filters: Optional[Dict] = None, operation: str = None, name: Optional[str] = None, description: Optional[str] = None, join_name: Optional[str] = None, page: Optional[int] = None, page_size: Optional[int] = None, total_pages: Optional[int] = None) -> io.BytesIO:
    """Formatea datos de /api/v1/join/predefined para exportación a Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Consulta Predefinida"

    # Estilos
    title_font = Font(bold=True, size=20, name='Helvetica')
    body_font = Font(bold=True, size=12, name='Helvetica')
    meta_font = Font(size=10, name='Helvetica')
    header_font = Font(bold=True, size=10, color="FFFFFF", name='Helvetica')
    data_font = Font(size=9, name='Helvetica')
    header_fill = PatternFill(start_color="003087", end_color="003087", fill_type="solid")  # Azul oscuro
    data_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gris claro
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Título y metadatos
    ws.append(["Reporte de Consulta Predefinida"])
    ws.merge_cells('A1:G1')
    ws['A1'].font = title_font
    ws['A1'].alignment = align_center
    ws.append(["Generado", datetime.now(timezone.utc).strftime('%Y-%m-%d')])
    ws['A2'].font = body_font
    ws['B2'].font = body_font
    if operation:
        ws.append(["Operación", operation.capitalize()])
        ws['A3'].font = body_font
        ws['B3'].font = body_font
    row_offset = 4 if operation else 3
    if filters:
        exclude_keys = {"format", "save_report", "name", "description"}
        used_filters = {k: v for k, v in filters.items() if v not in [None, "", [], {}] and k not in exclude_keys}
        if used_filters:
            ws.append(["Filtros", ""])
            ws['A' + str(row_offset)].font = body_font
            for key, value in used_filters.items():
                pretty_value = json.dumps(value, ensure_ascii=False, indent=2) if isinstance(value, (dict, list)) else str(value)
                ws.append([f"• {key}", pretty_value])
                ws['A' + str(row_offset + 1)].font = meta_font
                ws['B' + str(row_offset + 1)].font = meta_font
                row_offset += 1
            ws.append([])
            row_offset += 1
    ws.append(["Nombre", name if name else "Sin nombre"])
    ws['A' + str(row_offset)].font = body_font
    ws['B' + str(row_offset)].font = body_font
    row_offset += 1
    ws.append(["Descripción", description if description else "Sin descripción"])
    ws['A' + str(row_offset)].font = body_font
    ws['B' + str(row_offset)].font = body_font
    row_offset += 1
    ws.append(["Consulta Predefinida", join_name if join_name else "No especificada"])
    ws['A' + str(row_offset)].font = body_font
    ws['B' + str(row_offset)].font = body_font
    ws.append([])  # Espacio
    row_offset += 2

    # Paginación
    if page and page_size and total_pages:
        ws.append(["Paginación"])
        ws.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=3)
        ws['A' + str(row_offset)].font = body_font
        ws['A' + str(row_offset)].alignment = align_center
        row_offset += 1
        ws.append(["Página", "Tamaño de Página", "Total de Páginas"])
        for cell in ws[row_offset]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = align_center
        row_offset += 1
        ws.append([str(page), str(page_size), str(total_pages)])
        for cell in ws[row_offset]:
            cell.font = data_font
            cell.fill = data_fill
            cell.border = border
            cell.alignment = align_center
        row_offset += 2

    # Procesar datos
    if data:
        headers = list(data[0].keys()) if data else []
        headers_display = [key.replace("_", " ").title() for key in headers]
        ws.append(["Resultados"])
        ws.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=len(headers))
        ws['A' + str(row_offset)].font = body_font
        ws['A' + str(row_offset)].alignment = align_center
        row_offset += 1
        ws.append(headers_display)
        for cell in ws[row_offset]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = align_center
        row_offset += 1
        for item in data:
            row = []
            for key in headers:
                value = item.get(key, "")
                if isinstance(value, (int, float)) and key in ["total", "subtotal", "unit_value", "amount", "concept_amount", "tax_amount", "total_tax_amount", "total_amount", "avg_amount", "payment_amount"]:
                    row.append(f"${float(value):.2f}")
                else:
                    row.append(str(value) if value is not None else "")
            ws.append(row)
            for cell in ws[row_offset]:
                cell.font = data_font
                cell.fill = data_fill
                cell.border = border
                cell.alignment = align_center
            row_offset += 1

        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
    else:
        ws.append(["No hay datos disponibles"])
        ws['A' + str(ws.max_row)].font = meta_font
        ws['A' + str(ws.max_row)].alignment = align_center

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def format_predefined_join_pdf(data: list, filters: Optional[Dict] = None, operation: str = None, name: Optional[str] = None, description: Optional[str] = None, join_name: Optional[str] = None, page: Optional[int] = None, page_size: Optional[int] = None, total_pages: Optional[int] = None) -> io.BytesIO:
    """Formatea datos de /api/v1/join/predefined para exportación a PDF."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, rightMargin=inch/2, leftMargin=inch/2, topMargin=inch, bottomMargin=inch)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='Title', fontSize=20, leading=20, spaceAfter=20, alignment=1, fontName='Helvetica-Bold')
    meta_style = ParagraphStyle(name='Meta', fontSize=10, leading=12, spaceAfter=10, fontName='Helvetica')
    subtitle_style = ParagraphStyle(name='Subtitle', fontSize=12, leading=14, spaceAfter=8, fontName='Helvetica-Bold', alignment=1)
    body_style = ParagraphStyle(name='Body', fontSize=12, leading=14, spaceAfter=9, fontName='Helvetica-Bold')

    # Título y metadatos
    elements.append(Paragraph("Reporte de Consulta Predefinida", title_style))
    elements.append(Paragraph(f"Generado: {datetime.now(timezone.utc).strftime('%Y-%m-%d')}", body_style))
    if operation:
        elements.append(Paragraph(f"Operación: {operation.capitalize()}", body_style))
    if filters:
        exclude_keys = {"format", "save_report", "name", "description"}
        used_filters = {k: v for k, v in filters.items() if v not in [None, "", [], {}] and k not in exclude_keys}
        if used_filters:
            elements.append(Paragraph("Filtros:", body_style))
            for key, value in used_filters.items():
                pretty_value = json.dumps(value, ensure_ascii=False, indent=2) if isinstance(value, (dict, list)) else str(value)
                elements.append(Paragraph(f"&#8226; {key}: {pretty_value}", meta_style))
    elements.append(Paragraph(f"Nombre: {name if name else 'Sin nombre'}", body_style))
    elements.append(Paragraph(f"Descripción: {description if description else 'Sin descripción'}", body_style))
    elements.append(Paragraph(f"Consulta Predefinida: {join_name if join_name else 'No especificada'}", body_style))
    elements.append(Spacer(1, 0.3 * inch))

    # Paginación
    if page and page_size and total_pages:
        pagination_headers = ["Página", "Tamaño de Página", "Total de Páginas"]
        pagination_row = [str(page), str(page_size), str(total_pages)]
        pagination_table = Table([pagination_headers, pagination_row], colWidths=[2 * inch] * 3)
        pagination_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('WORDWRAP', (0, 0), (-1, -1)),
        ]))
        elements.append(Paragraph("Paginación", subtitle_style))
        elements.append(pagination_table)
        elements.append(Spacer(1, 0.3 * inch))

    # Procesar datos
    if data:
        headers = list(data[0].keys())
        headers_display = [key.replace("_", " ").title() for key in headers]
        table_data = [headers_display]
        for item in data:
            row = []
            for key in headers:
                value = item.get(key, "")
                if isinstance(value, (int, float)) and key in ["total", "subtotal", "unit_value", "amount", "concept_amount", "tax_amount", "total_tax_amount", "total_amount", "avg_amount", "payment_amount"]:
                    row.append(f"${float(value):.2f}")
                else:
                    row.append(str(value) if value is not None else "")
            table_data.append(row)

        # Calcular anchos de columna dinámicamente
        col_widths = [min(7.0 / len(headers), 2.0) * inch for _ in headers]
        data_table = Table(table_data, colWidths=col_widths)
        data_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('WORDWRAP', (0, 0), (-1, -1)),
        ]))
        elements.append(Paragraph("Resultados", subtitle_style))
        elements.append(data_table)
    else:
        elements.append(Paragraph("No hay datos disponibles", styles['Normal']))

    doc.build(elements)
    output.seek(0)
    return output

async def generate_report_from_data(
    data: Any,
    format_type: str,
    user: Dict,
    cfdi_id: Optional[int] = None,
    save_report: bool = False,
    name: Optional[str] = None,
    description: Optional[str] = None,
    filters: Optional[Dict] = None,
    operation: Optional[str] = None,
    join_name: Optional[str] = None,
    page: Optional[int] = None,
    page_size: Optional[int] = None,
    total_pages: Optional[int] = None
) -> Dict:
    """
    Genera un reporte a partir de datos proporcionados en el formato especificado (json, xml, csv, excel, pdf).
    Opcionalmente guarda el reporte en la base de datos y publica un evento.

    Args:
        data: Datos a convertir en reporte (se aceptan listas, diccionarios o valores simples).
        format_type: Formato del reporte (json, xml, csv, excel, pdf).
        user: Información del usuario (incluye rfc).
        cfdi_id: ID del CFDI asociado (opcional).
        save_report: Indica si se debe guardar el reporte en la DB.
        name: Nombre del reporte (opcional).
        description: Descripción del reporte (opcional).
        filters: Filtros utilizados en la consulta (opcional).
        operation: Operación ejecutada (e.g., predefined_join).
        join_name: Nombre de la consulta predefinida (opcional).
        page: Número de página actual (opcional).
        page_size: Tamaño de la página (opcional).
        total_pages: Total de páginas (opcional).

    Returns:
        Dict: Contiene el contenido del reporte (Base64 para excel/pdf, texto para json/xml/csv),
              tipo de contenido y ID del reporte (si se guarda).
    """
    start_time = datetime.now(timezone.utc)
    format_type = format_type.lower()

    # Validar formato
    if format_type not in ["json", "xml", "csv", "excel", "pdf"]:
        logger.error(f"Formato inválido recibido: {format_type}")
        raise HTTPException(status_code=400, detail="Invalid format. Use json, xml, csv, excel, or pdf")

    # Normalizar datos de entrada a una lista
    if data is None:
        logger.warning("Datos nulos recibidos para la generación del reporte")
        data = []
    elif not isinstance(data, list):
        logger.info(f"Convirtiendo datos no lista a lista: tipo recibido {type(data)}")
        data = [data] if isinstance(data, dict) else [{"value": data}]

    # Validar que los elementos sean procesables
    if data and not all(isinstance(item, dict) for item in data):
        logger.error(f"Datos inválidos: todos los elementos deben ser diccionarios, recibido: {data}")
        raise HTTPException(status_code=400, detail=f"Invalid data: all items must be dictionaries, got {type(data[0]) if data else None}")

    # Definir tipo de contenido
    content_type = {
        "json": "application/json",
        "xml": "application/xml",
        "csv": "text/csv",
        "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "pdf": "application/pdf"
    }[format_type]

    # Seleccionar función de formato específica
    format_functions = {
        "predefined_join": {
            "excel": format_predefined_join_excel,
            "pdf": format_predefined_join_pdf
        }
    }

    file_content = None
    file_content_base64 = None

    try:
        if format_type in ["json", "xml", "csv"]:
            if format_type == "json":
                file_content = json.dumps(data, ensure_ascii=False, indent=2)
                file_content_base64 = file_content
            elif format_type == "xml":
                file_content = dicttoxml(data, custom_root="data", attr_type=False).decode("utf-8")
                file_content_base64 = file_content
            elif format_type == "csv":
                flattened_data = [flatten_dict(item) for item in data]
                df = pd.DataFrame(flattened_data if data else [])
                file_content = df.to_csv(index=False, encoding='utf-8') if not df.empty else ""
                file_content_base64 = file_content
        else:
            # Usar función de formato específica según la operación
            format_func = format_functions.get(operation, {}).get(format_type)
            if not format_func:
                logger.warning(f"No se encontró función de formato para operación {operation} y formato {format_type}. Usando formato genérico.")
                # Fallback genérico para excel
                if format_type == "excel":
                    flattened_data = [flatten_dict(item) for item in data]
                    df = pd.DataFrame(flattened_data if data else [])
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        df.to_excel(writer, index=False)
                    file_content = output.getvalue()
                    file_content_base64 = base64.b64encode(file_content).decode("utf-8")
                    output.close()
                # Fallback genérico para pdf
                elif format_type == "pdf":
                    output = io.BytesIO()
                    doc = SimpleDocTemplate(output, pagesize=letter)
                    elements = []
                    if data:
                        table_data = [list(data[0].keys())]
                        for row in data:
                            table_data.append([str(value) if value is not None else "" for value in row.values()])
                        table = Table(table_data)
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ]))
                        elements.append(table)
                    else:
                        styles = getSampleStyleSheet()
                        elements.append(Paragraph("No hay datos disponibles", styles['Normal']))
                    doc.build(elements)
                    file_content = output.getvalue()
                    file_content_base64 = base64.b64encode(file_content).decode("utf-8")
                    output.close()
            else:
                output = format_func(data, filters, operation, name, description, join_name, page, page_size, total_pages)
                file_content = output.getvalue()
                file_content_base64 = base64.b64encode(file_content).decode("utf-8")
                output.close()

    except Exception as e:
        logger.error(f"Error generando reporte en formato {format_type}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate {format_type} report: {str(e)}")

    # Guardar el reporte en la base de datos si se solicita
    report_id = None
    if save_report:
        try:
            report_data = {
                "format": format_type.upper(),
                "file_content": file_content_base64,
                "user_id": user["rfc"],
                "name": name,
                "description": description,
                "filters": json.dumps(filters) if filters else None,
                "operation": operation
            }
            if cfdi_id is not None:
                report_data["cfdi_id"] = cfdi_id

            report = await db.report.create(data=report_data)
            report_id = report.id

            # Publicar evento de reporte generado
            await publish_event("report_generated", {
                "report_id": report_id,
                "user_id": user["rfc"],
                "format": format_type,
                "timestamp": datetime.now(timezone.utc).isoformat()
            })

            logger.info(f"Reporte guardado con ID: {report_id} para RFC: {user['rfc']} en formato {format_type}")
        except Exception as e:
            logger.error(f"Error al guardar el reporte: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Failed to save report: {str(e)}")

    logger.info(
        f"Reporte generado (formato: {format_type}, guardado: {save_report}, "
        f"nombre: {name}, descripción: {description}, operación: {operation}, "
        f"consulta: {join_name}) para RFC: {user['rfc']} en "
        f"{(datetime.now(timezone.utc) - start_time).total_seconds():.2f} segundos"
    )
    return {
        "content": file_content_base64 if format_type in ["excel", "pdf"] else file_content,
        "content_type": content_type,
        "report_id": report_id
    }